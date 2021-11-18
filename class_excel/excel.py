# -*- coding: utf-8 -*-
# @Time : 2021/11/18 15:09
# @Author : Limusen
# @File : excel


import os
from openpyxl import load_workbook

"""

1.准备测试数据
2.load_workbook模块,打开测试数据文件,生成WorkBook对象(wb)
3.根据表单名称选择表单(sh): wb["表单名称"]
4.在表单当中,获取单元格的数据 列
    4.1 单元格对象: sh.cell(row,column) 下标从1开始
    4.2 .value获取单元格的值

5.得到当前表单当中,总行数和总列数
    sh.max_row # 总行数
    sh.max_column # 总列数

6.修改单元格数据
    6.1 修改单元格数据: sh.cell(row,column).value = 新的值
    
7.修改单元格之后，需要保存工作簿
    7.1 Workbook().save(工作路径)

"""

file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'login_cases.xlsx')
print(file_path)

# 　1.加载excel数据文件
wb = load_workbook(file_path)
# 　2.根据表单名称选择表单
sh = wb['login']
print(sh)
#  3.单元格对象: sh.cell(row,column)
cel = sh.cell(2, 2)
print(cel.value)

# 5.1获取总行数
print(sh.max_row)
# 5.2获取总列数
print(sh.max_column)

print("=========================")
# 6.3 修改单元格数据: sh.cell(row,column).value = 新的值
sh.cell(2, 2).value = "123456"
print(sh.cell(2, 2).value)

# 7.1 保存工作蒲
# 　WorkBook对象(wb).save(地址)
# wb.save("save_as_another_excel.xlsx") # 另存为
wb.save(file_path)